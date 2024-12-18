from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
import os

app = Flask(__name__)
CORS(app)  # Permite todas as origens por padrão

EXCEL_FILE_PATH = "os.getenv("EXCEL_FILE_PATH", "compromissos.xlsx")"

def create_excel():
    if not os.path.exists(EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Compromissos"
        headers = ["ID", "Nome", "Data", "Hora", "Descrição"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        wb.save(EXCEL_FILE_PATH)

def add_compromisso(id, nome, data, hora, descricao):
    if os.path.exists(EXCEL_FILE_PATH):
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
    else:
        create_excel()
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active

    ws.append([id, nome, data, hora, descricao])
    wb.save(EXCEL_FILE_PATH)

def generate_id():
    return int(datetime.now().timestamp())  # Gerando ID único baseado no timestamp

@app.route('/agendar', methods=['POST'])
def agendar():
    data = request.get_json()
    print(data)  # Debug: Verifique o conteúdo do JSON recebido
    # Acesse os dados corretamente
    nome = data['nomePaciente']
    cpf = data['cpf']
    data_nascimento = data['dataNascimento']
    procedimento = data['procedimento']
    status = data['status']

    create_excel()
    compromisso_id = generate_id()
    add_compromisso(compromisso_id, nome, data_nascimento, hora, procedimento)
    
    return jsonify({"message": "Compromisso agendado com sucesso!", "compromisso": data}), 200

if __name__ == '__main__':
    app.run(debug=True)
